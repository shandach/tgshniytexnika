"""GET /api/export — скачивание XLSX-отчёта по заявкам."""

from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.database import get_session
from app.models.request import Request
from app.models.user import User

router = APIRouter(prefix="/export", tags=["export"])


STATUS_LABELS = {
    "new": "Новая", "in_progress": "В обработке", "closed": "Закрыта",
}
DECISION_LABELS = {
    "pending": "На рассмотрении", "approved": "Одобрено", "rejected": "Отклонено",
}
TYPE_LABELS = {
    "replacement": "Замена", "new_issue": "Выдача", "repair": "Поломка",
}


@router.get("")
async def export_xlsx(
    branch: Optional[str] = None,
    type: Optional[str] = None,
    date_from: Optional[str] = Query(None, alias="from"),
    date_to: Optional[str] = Query(None, alias="to"),
    session: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Генерирует XLSX-файл с заявками по фильтрам."""
    try:
        import openpyxl
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl не установлен. pip install openpyxl")

    stmt = select(Request).order_by(Request.created_at.desc())
    filters = []
    if branch:
        filters.append(Request.bhm_code_snapshot == branch)
    if type:
        type_map = {"replacement": "replacement", "new": "new_issue", "repair": "repair"}
        filters.append(Request.request_type == type_map.get(type, type))
    if date_from:
        filters.append(func.date(Request.created_at) >= date_from)
    if date_to:
        filters.append(func.date(Request.created_at) <= date_to)
    if filters:
        stmt = stmt.where(*filters)

    rows = (await session.scalars(stmt)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["№ заявки", "Филиал", "BXM код", "Тип", "Оборудование",
               "Инвентарь", "Сотрудник", "Статус", "Решение", "Дата создания"]
    ws.append(headers)

    # Стилизация хедеров
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3525CD", end_color="3525CD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for r in rows:
        ws.append([
            r.request_number,
            r.branch_name_snapshot,
            r.bhm_code_snapshot,
            TYPE_LABELS.get(r.request_type, r.request_type),
            "Принтер" if r.equipment_type == "printer" else "Компьютер",
            r.inventory_code_snapshot or "—",
            r.employee_fio_snapshot,
            STATUS_LABELS.get(r.status, r.status),
            DECISION_LABELS.get(r.final_decision, r.final_decision),
            r.created_at.strftime("%d.%m.%Y %H:%M") if r.created_at else "",
        ])

    # Авто-ширина колонок
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=report.xlsx"},
    )
